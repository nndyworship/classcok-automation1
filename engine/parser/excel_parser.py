"""
강사별 시트에서 월별 주차 데이터를 추출하고 촉감놀이를 감지한다.

시트 구조 (행 1: 제목, 행 2: 헤더, 행 3~: 데이터):
  A: 학기  B: 주차(예: "3월 1주")  C: 수업 주제  D: 활동 내용  E: 준비물
"""

from __future__ import annotations
from dataclasses import dataclass, field
from pathlib import Path
import openpyxl

from engine.parser.sensory_detector import detect as detect_sensory, SensoryResult

# 학기 → 월 매핑
SEMESTER_MONTHS: dict[str, list[int]] = {
    "봄학기": [3, 4, 5],
    "여름학기": [6, 7, 8],
    "가을학기": [9, 10, 11],
    "겨울학기": [12, 1, 2],
}

COL_SEMESTER = 1  # A
COL_WEEK_LABEL = 2  # B  "3월 1주"
COL_TITLE = 3      # C
COL_CONTENT = 4    # D
COL_SUPPLIES = 5   # E
DATA_START_ROW = 3


@dataclass
class WeekData:
    week: int
    title: str
    content: str
    supplies: str
    sensory: SensoryResult


@dataclass
class MonthData:
    instructor: str
    year: int
    month: int
    weeks: list[WeekData] = field(default_factory=list)


def _month_from_label(label: str) -> int | None:
    """'3월 1주' → 3"""
    if not label:
        return None
    try:
        return int(label.split("월")[0].strip())
    except (ValueError, IndexError):
        return None


def _week_num_from_label(label: str) -> int | None:
    """'3월 1주' → 1"""
    if not label:
        return None
    try:
        return int(label.split("주")[0].split()[-1].strip())
    except (ValueError, IndexError):
        return None


def parse(
    excel_path: str | Path,
    instructor_sheet: str,
    target_month: int,
    year: int = 2026,
) -> MonthData:
    """
    엑셀 파일에서 특정 강사의 특정 월 주차 데이터를 추출한다.

    Args:
        excel_path: 엑셀 파일 경로
        instructor_sheet: 시트명 (예: "최보라T")
        target_month: 추출할 월 (1~12)
        year: 연도 (기본 2026)

    Returns:
        MonthData (weeks 리스트 포함)

    Raises:
        KeyError: 시트가 존재하지 않을 때
        ValueError: 해당 월 데이터가 없을 때
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)

    if instructor_sheet not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise KeyError(f"시트 '{instructor_sheet}' 없음. 사용 가능: {available}")

    ws = wb[instructor_sheet]
    result = MonthData(instructor=instructor_sheet, year=year, month=target_month)

    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        week_label = str(row[COL_WEEK_LABEL - 1] or "").strip()
        month = _month_from_label(week_label)
        if month != target_month:
            continue

        week_num = _week_num_from_label(week_label)
        title = str(row[COL_TITLE - 1] or "").strip()
        content = str(row[COL_CONTENT - 1] or "").strip()
        base_supplies = str(row[COL_SUPPLIES - 1] or "").strip()

        sensory = detect_sensory(content, base_supplies or "편안한 복장, 물티슈, 마실물")

        result.weeks.append(WeekData(
            week=week_num or len(result.weeks) + 1,
            title=title,
            content=content,
            supplies=sensory.supplies,
            sensory=sensory,
        ))

    wb.close()

    if not result.weeks:
        raise ValueError(f"'{instructor_sheet}' 시트에서 {target_month}월 데이터를 찾을 수 없음")

    result.weeks.sort(key=lambda w: w.week)
    return result


def list_instructors(excel_path: str | Path) -> list[str]:
    """엑셀 파일의 강사 시트 목록을 반환한다 (스케줄·프린트 시트 제외)."""
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    excluded = {"스케줄", "프린트"}
    sheets = [s for s in wb.sheetnames if s not in excluded]
    wb.close()
    return sheets
